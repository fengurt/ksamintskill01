#!/usr/bin/env python3
"""roundtrip_xlsx — prove (or disprove) that a conversion of a workbook was lossless.

Requires openpyxl.

    python roundtrip_xlsx.py extract <original.xlsx> --out cells.ndjson [--layout layout.json]
    python roundtrip_xlsx.py compare <original.xlsx> <converted> [--sheet NAME]

<converted> may be .xlsx, .ndjson (as produced by `extract`), or .csv (one sheet;
pass --sheet to say which sheet of the original it came from).

The verdict is reported per level:
    L1_values    every non-empty cell's cached value matches
    L2_formulas  every formula string matches            (xlsx / ndjson only)
    L3_layout    merged ranges, hidden rows/cols, number formats match (xlsx only)
Anything not checkable on the given target is reported as "not_checkable",
never as "pass" — silence is not evidence.
"""
from __future__ import annotations

import argparse
import csv
import datetime as _dt
import json
import sys

try:
    import openpyxl  # type: ignore
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


def norm(v):
    """Normalise a cell value for comparison across formats."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return int(v)
        return round(v, 10)
    if isinstance(v, str):
        s = v.strip()
        # numbers that came back from CSV as text
        try:
            f = float(s)
            return norm(f)
        except ValueError:
            pass
        if s.upper() in ("TRUE", "FALSE"):
            return s.upper() == "TRUE"
        # dates printed by a CSV writer ("2026-01-09 00:00:00", "2026-01-09")
        try:
            return _dt.datetime.fromisoformat(s.replace(" ", "T")).isoformat()
        except ValueError:
            return s
    return v


def read_xlsx(path: str) -> dict:
    """Returns {sheet: {'values': {coord: v}, 'formulas': {coord: f}, 'layout': {...}}}"""
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb_f.worksheets:
        wv = wb_v[ws.title]
        values, formulas, numfmt = {}, {}, {}
        uncached = []
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cached = wv[c.coordinate].value
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas[c.coordinate] = c.value
                    if cached is None:
                        # formula never calculated (saved by a library, or calc off):
                        # there is no value to compare, so it cannot count for or against L1
                        uncached.append(c.coordinate)
                        continue
                values[c.coordinate] = norm(cached)
                if c.number_format and c.number_format != "General":
                    numfmt[c.coordinate] = c.number_format
        layout = {
            "merged": sorted(str(r) for r in ws.merged_cells.ranges),
            "hidden_rows": sorted(i for i, d in ws.row_dimensions.items() if d.hidden),
            "hidden_cols": sorted(k for k, d in ws.column_dimensions.items() if d.hidden),
            "number_formats": numfmt,
            "sheet_state": ws.sheet_state,
        }
        out[ws.title] = {"values": values, "formulas": formulas, "layout": layout, "uncached_formula_cells": uncached}
    return out


def read_ndjson(path: str) -> dict:
    out = {}
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            if not line.strip():
                continue
            r = json.loads(line)
            s = out.setdefault(r["sheet"], {"values": {}, "formulas": {}, "layout": None})
            if r.get("formula"):
                s["formulas"][r["coord"]] = r["formula"]
                if r.get("value") is None:
                    continue
            s["values"][r["coord"]] = norm(r.get("value"))
    return out


def read_csv_as_sheet(path: str, sheet_name: str) -> dict:
    values = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        for ri, row in enumerate(csv.reader(fh), start=1):
            for ci, cell in enumerate(row, start=1):
                if cell == "":
                    continue
                values[f"{openpyxl.utils.get_column_letter(ci)}{ri}"] = norm(cell)
    return {sheet_name: {"values": values, "formulas": {}, "layout": None}}


def extract(path: str, out: str, layout_out: str | None):
    data = read_xlsx(path)
    n = 0
    with open(out, "w", encoding="utf-8") as fh:
        for sheet, s in data.items():
            coords = list(s["values"]) + [c for c in s.get("uncached_formula_cells", []) if c not in s["values"]]
            for coord in coords:
                v = s["values"].get(coord)
                rec = {"sheet": sheet, "coord": coord, "value": v}
                if coord in s["formulas"]:
                    rec["formula"] = s["formulas"][coord]
                if coord in s["layout"]["number_formats"]:
                    rec["number_format"] = s["layout"]["number_formats"][coord]
                fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
                n += 1
    print(f"wrote {n} cells to {out}")
    if layout_out:
        with open(layout_out, "w", encoding="utf-8") as fh:
            json.dump({k: v["layout"] for k, v in data.items()}, fh, ensure_ascii=False, indent=2)
        print(f"wrote layout to {layout_out}")


def diff_dict(a: dict, b: dict, limit: int = 20):
    only_a = sorted(set(a) - set(b))
    only_b = sorted(set(b) - set(a))
    changed = sorted(k for k in set(a) & set(b) if a[k] != b[k])
    return {
        "missing_in_converted": len(only_a),
        "extra_in_converted": len(only_b),
        "changed": len(changed),
        "examples": [{"cell": k, "original": a[k], "converted": b[k]} for k in changed[:limit]]
                    + [{"cell": k, "original": a[k], "converted": None} for k in only_a[:limit]],
    }


def compare(original: str, converted: str, sheet: str | None):
    src = read_xlsx(original)
    if converted.lower().endswith((".xlsx", ".xlsm")):
        tgt = read_xlsx(converted)
        can_formulas, can_layout = True, True
    elif converted.lower().endswith((".ndjson", ".jsonl")):
        tgt = read_ndjson(converted)
        can_formulas, can_layout = True, False
    elif converted.lower().endswith((".csv", ".tsv")):
        if sheet is None:
            if len(src) == 1:
                sheet = next(iter(src))
            else:
                sys.exit("--sheet is required when the original has several sheets")
        tgt = read_csv_as_sheet(converted, sheet)
        src = {sheet: src[sheet]}
        can_formulas, can_layout = False, False
    else:
        sys.exit("unsupported converted format")

    report = {"original": original, "converted": converted, "sheets": {}, "verdict": {}}
    l1 = l2 = l3 = True
    for name, s in src.items():
        t = tgt.get(name)
        if t is None:
            report["sheets"][name] = {"status": "sheet missing in converted"}
            l1 = l2 = l3 = False
            continue
        v = diff_dict(s["values"], t["values"])
        entry = {"L1_values": v}
        if s.get("uncached_formula_cells"):
            entry["L1_note"] = (f"{len(s['uncached_formula_cells'])} formula cells have no cached value in the original "
                                f"(never calculated); L1 was checked on the remaining cells only")
        if v["missing_in_converted"] or v["changed"]:
            l1 = False
        if can_formulas:
            f = diff_dict(s["formulas"], t["formulas"])
            entry["L2_formulas"] = f
            if f["missing_in_converted"] or f["changed"]:
                l2 = False
        if can_layout and t["layout"]:
            lay = {}
            for k in ("merged", "hidden_rows", "hidden_cols", "sheet_state"):
                lay[k] = "same" if s["layout"][k] == t["layout"][k] else {"original": s["layout"][k], "converted": t["layout"][k]}
            nf = diff_dict(s["layout"]["number_formats"], t["layout"]["number_formats"], limit=10)
            lay["number_formats"] = nf
            entry["L3_layout"] = lay
            if any(x != "same" for k, x in lay.items() if k != "number_formats") or nf["missing_in_converted"] or nf["changed"]:
                l3 = False
        report["sheets"][name] = entry
    extra_sheets = sorted(set(tgt) - set(src))
    if extra_sheets:
        report["extra_sheets_in_converted"] = extra_sheets
    report["verdict"] = {
        "L1_values": "pass" if l1 else "fail",
        "L2_formulas": ("pass" if l2 else "fail") if can_formulas else "not_checkable",
        "L3_layout": ("pass" if l3 else "fail") if can_layout else "not_checkable",
        "L4_objects": "not_checkable",
        "L5_code_and_links": "not_checkable",
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return report


def main(argv=None):
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    e = sub.add_parser("extract")
    e.add_argument("original")
    e.add_argument("--out", required=True)
    e.add_argument("--layout")
    c = sub.add_parser("compare")
    c.add_argument("original")
    c.add_argument("converted")
    c.add_argument("--sheet")
    args = ap.parse_args(argv)
    if args.cmd == "extract":
        extract(args.original, args.out, args.layout)
    else:
        compare(args.original, args.converted, args.sheet)


if __name__ == "__main__":
    main()
